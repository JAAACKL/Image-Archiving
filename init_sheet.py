import os, sys
import openpyxl 

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = 'Thumbnail'
ws['B1'] = 'Filename'
ws['C1'] = 'Format'
ws['D1'] = 'Dimensions'
ws['E1'] = 'File Size (kb)'
ws['F1'] = 'Compressed Filename'
ws['G1'] = 'New File Size'
ws['H1'] = 'Size Reduction'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['F'].width = 30

wb.save('sample.xlsx')