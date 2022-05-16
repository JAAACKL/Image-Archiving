import os, sys
from PIL import Image
import openpyxl, zlib

wb = openpyxl.load_workbook("sample.xlsx") 
ws = wb.active
row = max(len(ws['A']), 2)

# dimensions for thumbnail picture
thumbnail_size = (128, 128)

for infile in sys.argv[1:]:
    ws.row_dimensions[row].height = 100
    archive = infile + '.bin'
    try:
        print('processing: ' + infile)
        with Image.open(infile) as im:
            ws['B' + str(row)] = im.filename
            ws['C' + str(row)] = im.format
            ws['D' + str(row)] = str(im.size)
            ws['E' + str(row)] = os.path.getsize(infile) / 1000
            ws['E' + str(row)].number_format = '0.00"kb"'
            
            outfile = os.path.splitext(infile)[0] + ".thumbnail"
            if not os.path.exists(outfile) and infile != outfile:
                try:
                    rgb_im = im.convert('RGB')
                    rgb_im.thumbnail(thumbnail_size, Image.ANTIALIAS)
                    rgb_im.save(outfile, "WEBP")
                except OSError:
                    print("cannot create thumbnail for", infile)
            img = openpyxl.drawing.image.Image(outfile)
            img.anchor = 'A' + str(row)
            ws.add_image(img)
            compressed_file_name = 'compressed_' + infile + '.webp'
            im.save(compressed_file_name,optimize=True,lossless=(im.format=='PNG' or im.format=='GIF'),format="WEBP")
            smaller_file = compressed_file_name
            if os.path.getsize(compressed_file_name) >= os.path.getsize(infile):
                smaller_file = infile
            with open(smaller_file, "rb") as cim:
                compressed_data = zlib.compress(cim.read(), level=9)
                f = open(archive, "wb")
                f.write(compressed_data)
                f.close()
                os.remove(compressed_file_name)
            ws['F' + str(row)] = archive
            ws['G' + str(row)] = os.path.getsize(archive) / 1000
            ws['G' + str(row)].number_format = '0.00"kb"'
            ws['H' + str(row)] = float(ws['E' + str(row)].value) - float(ws['G' + str(row)].value)
            ws['H' + str(row)].number_format = '0.00"kb"'
    except Exception as e:
        print(e)
        ws['B' + str(row)] = infile
        ws['C' + str(row)] = os.path.splitext(infile)[-1][1:]
        ws['D' + str(row)] = '/'
        ws['E' + str(row)] = os.path.getsize(infile) / 1000
        ws['E' + str(row)].number_format = '0.00"kb"'
        with open(infile, "rb") as cim:
            compressed_data = zlib.compress(cim.read(), level=9)
            f = open(archive, "wb")
            f.write(compressed_data)
            f.close()
        ws['F' + str(row)] = archive
        ws['G' + str(row)] = os.path.getsize(archive) / 1000
        ws['G' + str(row)].number_format = '0.00"kb"'
        ws['H' + str(row)] = float(ws['E' + str(row)].value) - float(ws['G' + str(row)].value)
        ws['H' + str(row)].number_format = '0.00"kb"'
    row += 1

ws.row_dimensions[row].height = 50
ws['A' + str(row)] = 'Image Count:'
ws['B' + str(row)] = str(row - 2)
ws['C' + str(row)] = 'Disk\nSpace\nSaved:'
ws['D' + str(row)] = '= SUM(H2:H' + str(row - 1) + ')'
ws['E' + str(row)] = 'Previous\nTotal\nSize:'
ws['F' + str(row)] = '= SUM(E2:E' + str(row - 1) + ')'
ws['G' + str(row)] = 'Optimized\nPercentage:'
ws['H' + str(row)] = '= D' + str(row) + '/F' + str(row)
ws['H' + str(row)].number_format = '0%'
  
wb.save('sample.xlsx')